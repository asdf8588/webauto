"""
DataLoader - YAML测试数据加载器

从YAML文件加载测试用例数据
"""
import os
import yaml
import time
from pathlib import Path
from typing import List, Dict, Any, Optional


class DataLoader:
    """YAML测试数据加载器"""

    def __init__(self, data_dir: str = None):
        if data_dir is None:
            project_root = Path(__file__).parent.parent
            data_dir = project_root / "data"
        self.data_dir = Path(data_dir)

    def load(self, module: str, filename: str = "cases.yaml") -> List[Dict[str, Any]]:
        """
        加载指定模块的测试用例

        Args:
            module: 模块名 (如 "login", "user")
            filename: 文件名 (默认 cases.yaml)

        Returns:
            测试用例列表
        """
        file_path = self.data_dir / module / filename
        with open(file_path, 'r', encoding='utf-8') as f:
            cases = yaml.safe_load(f)
            return cases if cases else []

    def load_by_path(self, file_path: str) -> List[Dict[str, Any]]:
        """加载指定路径的YAML文件"""
        with open(file_path, 'r', encoding='utf-8') as f:
            return yaml.safe_load(f) or []

    @staticmethod
    def resolve_case(case: Dict[str, Any], context: Dict[str, Any] = None) -> Dict[str, Any]:
        """
        解析用例中的变量占位符

        支持:
        - {timestamp}: 当前时间戳
        - {user_id}: 上下文中的用户ID
        - {username}: 上下文中的用户名
        - {password}: 上下文中的密码
        """
        import copy
        resolved = copy.deepcopy(case)
        context = context or {}
        ts = str(int(time.time()))

        def replace(value):
            if isinstance(value, str):
                return (value
                    .replace("{timestamp}", ts)
                    .replace("{user_id}", str(context.get("user_id", "")))
                    .replace("{username}", context.get("username", ""))
                    .replace("{password}", context.get("password", ""))
                    .replace("{existing_user}", context.get("existing_user", "")))
            elif isinstance(value, dict):
                return {k: replace(v) for k, v in value.items()}
            elif isinstance(value, list):
                return [replace(item) for item in value]
            return value

        # 替换body中的变量
        if 'body' in resolved:
            resolved['body'] = replace(resolved['body'])

        # 替换url中的变量
        if 'url' in resolved:
            resolved['url'] = replace(resolved['url'])

        return resolved


# ═══════════════════════════════════════════════════════════════════════════
# 独立加载函数（兼容旧接口）
# ═══════════════════════════════════════════════════════════════════════════

def _resolve_data_file(filename: str) -> Path:
    """解析数据文件路径：优先相对项目根目录"""
    p = Path(filename)
    if p.is_absolute():
        return p
    # 尝试项目根目录 (utils 的父级)
    project_root = Path(__file__).parent.parent
    for candidate in [
        project_root / filename,
        project_root / "test_data" / filename,
        project_root / "data" / filename,
    ]:
        if candidate.exists():
            return candidate
    # 最后返回原始路径（让调用方收到有意义的 FileNotFoundError）
    return p


def load_excel(filename: str) -> Dict[str, List[Dict[str, Any]]]:
    """
    加载 Excel 文件（.xlsx / .xls），按 sheet 名分组返回。

    Args:
        filename: 文件名或相对/绝对路径
                  相对路径时依次查找：项目根目录、test_data/、data/

    Returns:
        {"sheet_name": [{col: value, ...}, ...], ...}
    """
    import openpyxl

    file_path = _resolve_data_file(filename)
    if not file_path.exists():
        raise FileNotFoundError(f"Excel 文件不存在: {file_path}")

    wb = openpyxl.load_workbook(file_path, data_only=True)
    result = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            result[sheet_name] = []
            continue
        headers = [
            str(h).strip() if h is not None else f"col_{i}"
            for i, h in enumerate(rows[0])
        ]
        data = []
        for row in rows[1:]:
            if all(v is None for v in row):
                continue
            record = {
                headers[i]: (row[i] if i < len(row) else None)
                for i in range(len(headers))
            }
            data.append(record)
        result[sheet_name] = data
    wb.close()
    return result


def load_yaml(filename: str) -> List[Dict[str, Any]]:
    """加载 YAML 文件（兼容旧接口）"""
    file_path = _resolve_data_file(filename)
    if not file_path.exists():
        raise FileNotFoundError(f"YAML 文件不存在: {file_path}")
    with open(file_path, "r", encoding="utf-8") as f:
        data = yaml.safe_load(f)
        return data if isinstance(data, list) else []
