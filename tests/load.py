# -*- coding: utf-8 -*-
"""
测试数据加载器
支持 Excel (.xlsx/.xls) 和 YAML 格式
"""
import os
import yaml
from pathlib import Path
from typing import List, Dict, Any

try:
    import openpyxl
except ImportError:
    openpyxl = None

try:
    import xlrd
except ImportError:
    xlrd = None


def load_excel(file_path: str) -> Dict[str, List[Dict[str, Any]]]:
    """
    加载 Excel 文件，按 sheet 名分组返回数据

    Args:
        file_path: Excel 文件路径（支持 .xlsx 和 .xls）

    Returns:
        {
            "sheet_name": [{col1: val1, col2: val2}, ...],
            ...
        }
    """
    file_path = Path(file_path)
    if not file_path.exists():
        raise FileNotFoundError(f"Excel 文件不存在: {file_path}")

    ext = file_path.suffix.lower()
    result = {}

    if ext == ".xlsx":
        if openpyxl is None:
            raise ImportError("请安装 openpyxl: pip install openpyxl")
        wb = openpyxl.load_workbook(file_path, data_only=True)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            result[sheet_name] = _parse_sheet(ws)
        wb.close()

    elif ext == ".xls":
        if xlrd is None:
            raise ImportError("请安装 xlrd: pip install xlrd")
        wb = xlrd.open_workbook(file_path)
        for idx in range(wb.nsheets):
            ws = wb.sheet_by_index(idx)
            sheet_name = ws.name
            rows = [ws.row_values(i) for i in range(ws.nrows)]
            result[sheet_name] = _parse_sheet_from_rows(rows)
        wb.release_resources()

    else:
        raise ValueError(f"不支持的格式: {ext}，仅支持 .xlsx 和 .xls")

    return result


def _parse_sheet(ws) -> List[Dict[str, Any]]:
    """解析 openpyxl Worksheet"""
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h).strip() if h is not None else f"col_{i}"
               for i, h in enumerate(rows[0])]
    data = []
    for row in rows[1:]:
        if all(v is None for v in row):
            continue
        record = {headers[i]: (row[i] if i < len(row) else None) for i in range(len(headers))}
        data.append(record)
    return data


def _parse_sheet_from_rows(rows: List) -> List[Dict[str, Any]]:
    """解析 xlrd rows"""
    if not rows:
        return []
    headers = [str(h).strip() if h else f"col_{i}" for i, h in enumerate(rows[0])]
    data = []
    for row in rows[1:]:
        if all(v == "" for v in row):
            continue
        record = {headers[i]: (row[i] if i < len(row) else "") for i in range(len(headers))}
        data.append(record)
    return data


def load_yaml(file_path: str) -> List[Dict[str, Any]]:
    """加载 YAML 文件"""
    with open(file_path, "r", encoding="utf-8") as f:
        data = yaml.safe_load(f)
        return data if isinstance(data, list) else []
